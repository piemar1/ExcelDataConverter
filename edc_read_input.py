# ! /usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl
from edc_sqlite import SQliteEdit
from string import ascii_uppercase


__author__ = 'Marcin Pieczyński'


class ReadInput(SQliteEdit):
    """
    The Class containing methods for reading data from excel fole.
    """
    def __init__(self, profile_name):
        """
        Inicjalization of ReadInput Class.
        """
        SQliteEdit.__init__(self)

        self.alfabet = ascii_uppercase

        self.daty = []                # a list with dates from input file
        self.cegly = []               # a list with CEGLA's names from input file
        self.cegly_position = {}      # a dictionary with CEGLA's names and positions from input file

        # Important data
        self.PnVa_data = {}           # a dictionary for PnVa data from input file
        self.UNITS_data = {}          # a dictionary for UNITS data from input file
        self.CEGLY_data = {}          # a dictionary for PnVa and UNITS data for CEGLA's from input file

        self.wb = None
        self.wb_sheets = None
        self.PnVa = None              # name of column with PnVa data
        self.UNITS = None             # name of column with UNITS data

        self.get_data_from_profile(profile_name)  # Reading data for profile 'profile_name' from Profile_database.db

    def open_input_file(self, input_file_path):
        """
        The method opens input excel file and reads names of the sheets
        """
        self.wb = openpyxl.load_workbook(input_file_path)
        self.wb_sheets = self.wb.get_sheet_names()

    def get_date(self):
        """
        The method for finding dates in in excel input files.
        """
        for sheet_name in self.wb_sheets:
            zakladka = self.wb.get_sheet_by_name(sheet_name)
            size = zakladka.calculate_dimension()            # return dimension of sheet as a string -> "A1:T60"

            for row in zakladka.iter_rows(size):
                for cell in row:
                    if "MNTH" in str(cell.value):
                        d = str(cell.value)[-7:]
                        if d not in self.daty:
                            self.daty.append(d)
                        if len(self.daty) == 3:
                            break

    def get_pnva_units_column(self):
        """
        The method for finding column with PnVa and Units data in input excel file.
        """
        for sheet_name in self.wb_sheets:
            zakladka = self.wb.get_sheet_by_name(sheet_name)
            size = zakladka.calculate_dimension()

            for row in zakladka.iter_rows(size):
                for cell in row:
                    if "Pn Va" in str(cell.value):
                        self.PnVa = cell.column
                    if "UNITS REPORT" in str(cell.value):
                        self.UNITS = cell.column
                    if self.PnVa and self.UNITS:
                        break

    def get_cegla_positions(self):
        """
        The method for finding names and positions of CEGLA's in input excel file.
        """
        for sheet_name in self.wb_sheets:
            zakladka = self.wb.get_sheet_by_name(sheet_name)
            no_of_row = zakladka.get_highest_row()              # returns numbers of row

            for r in range(no_of_row):                          # iteration for row
                wart = zakladka["A" + str(r+1)].value
                if wart and wart.strip() not in self.cegly:
                    self.cegly.append(wart.strip())

        for sheet_name in self.wb_sheets:
            positions = []

            self.cegly_position[sheet_name] = {str(k): [None, None] for k in self.cegly}  # creation of dictionary

            zakladka = self.wb.get_sheet_by_name(sheet_name)
            no_of_row = zakladka.get_highest_row()                  # returns numbers of row
            zakladka.iter_rows()

            for cegla in self.cegly:
                for r in range(no_of_row):                          # iteration for row
                    wart = zakladka["A" + str(r+1)].value
                    if wart and wart.strip() == cegla:
                        positions.append(r+1)

            for x in range(len(self.cegly)-1):              # saving ceglas positions data in dict
                self.cegly_position[sheet_name][self.cegly[x]][0] = positions[x]
                self.cegly_position[sheet_name][self.cegly[x]][1] = positions[x+1]-1

            self.cegly_position[sheet_name][self.cegly[-1]][0] = positions[-1]
            self.cegly_position[sheet_name][self.cegly[-1]][1] = int(no_of_row)

    def get_cegly_data(self):
        """
        The method for saving PnVa and UNITS data for CEGLA in a dictionary.
        """
        # Creation of dictionary for PnVa and UNITS data
        for cegla in self.cegly:
            self.CEGLY_data[cegla] = {lek: {"PnVa": [None, None, None], "UNITS": [None, None, None]}
                                      for lek in self.output_leki_cegly}

        # saving data in dictionary
        hit_number = len(self.output_leki_cegly)
        column_pnva_plus1 = self.alfabet[self.alfabet.index(self.PnVa)+1]
        column_pnva_plus2 = self.alfabet[self.alfabet.index(self.PnVa)+2]
        column_units_plus1 = self.alfabet[self.alfabet.index(self.UNITS)+1]
        column_units_plus2 = self.alfabet[self.alfabet.index(self.UNITS)+2]

        for sheet_name in self.wb_sheets:
            zakladka = self.wb.get_sheet_by_name(sheet_name)
            size = zakladka.calculate_dimension()

            for row in zakladka.iter_rows(size):
                for cell in row:
                    if str(cell.value).strip() in self.output_leki_cegly:
                        lek = str(cell.value).strip()
                        hit_number -= 1
                        wart1 = zakladka[str(self.PnVa) + str(cell.row)].value
                        wart2 = zakladka[str(column_pnva_plus1) + str(cell.row)].value
                        wart3 = zakladka[str(column_pnva_plus2) + str(cell.row)].value
                        wart4 = zakladka[str(self.UNITS) + str(cell.row)].value
                        wart5 = zakladka[str(column_units_plus1) + str(cell.row)].value
                        wart6 = zakladka[str(column_units_plus2) + str(cell.row)].value
                        for cegla in self.cegly:
                            x = self.cegly_position[sheet_name][cegla][0]
                            y = self.cegly_position[sheet_name][cegla][1]
                            if y >= cell.row >= x:
                                self.CEGLY_data[cegla][lek]["PnVa"] = [round(wart1, 2),
                                                                       round(wart2, 2),
                                                                       round(wart3, 2)]
                                self.CEGLY_data[cegla][lek]["UNITS"] = [round(wart4, 2),
                                                                        round(wart5, 2),
                                                                        round(wart6, 2)]
                                if hit_number == 0:
                                    break

    def get_pnva_units_data(self):
        """
        The method for saving PnVa and UNITS data from input excel file in two dictionary.
        """
        # Creation of dictionary for PnVa data
        for elem in range(len(self.output_zakladki)):
            self.PnVa_data[self.output_zakladki[elem]] = {str(k): {cegla: [None, None, None] for cegla in self.cegly}
                                                          for k in self.output_leki[elem]}

        # Creation of dictionary for UNITS data
        for elem in range(len(self.output_zakladki)):
            self.UNITS_data[self.output_zakladki[elem]] = {str(k): {cegla: [None, None, None] for cegla in self.cegly}
                                                           for k in self.output_leki[elem]}

        # Saving PnVa data in the dictionary
        [self.uzupelnianie_tabeli_pnva_units(self.output_leki[elem], self.output_zakladki[elem], self.PnVa)
         for elem in range(len(self.output_zakladki))]

        # Saving UNITS data in the dictionary
        [self.uzupelnianie_tabeli_pnva_units(self.output_leki[elem], self.output_zakladki[elem], self.UNITS)
         for elem in range(len(self.output_zakladki))]

    def uzupelnianie_tabeli_pnva_units(self, lista_lekow, nazwa_zakladki, rodzaj_danych):
        """
        The method for searching and saving PnVa and Units data from input file in dictionaries
        """
        if rodzaj_danych == self.PnVa:
            column_plus1 = self.alfabet[self.alfabet.index(self.PnVa)+1]
            column_plus2 = self.alfabet[self.alfabet.index(self.PnVa)+2]
            slownik = self.PnVa_data
        else:
            column_plus1 = self.alfabet[self.alfabet.index(self.UNITS)+1]
            column_plus2 = self.alfabet[self.alfabet.index(self.UNITS)+2]
            slownik = self.UNITS_data

        hit_number = len(lista_lekow)

        for sheet_name in self.wb_sheets:
            zakladka = self.wb.get_sheet_by_name(sheet_name)
            size = zakladka.calculate_dimension()

            for row in zakladka.iter_rows(size):
                for cell in row:
                    if str(cell.value).strip() in lista_lekow:
                        lek = str(cell.value).strip()
                        hit_number -= 1
                        wart1 = zakladka[str(rodzaj_danych) + str(cell.row)].value
                        wart2 = zakladka[str(column_plus1) + str(cell.row)].value
                        wart3 = zakladka[str(column_plus2) + str(cell.row)].value
                        for cegla in self.cegly:
                            x = self.cegly_position[sheet_name][cegla][0]
                            y = self.cegly_position[sheet_name][cegla][1]
                            if y >= cell.row >= x:
                                slownik[nazwa_zakladki][lek][cegla][0] = round(wart1, 2)
                                slownik[nazwa_zakladki][lek][cegla][1] = round(wart2, 2)
                                slownik[nazwa_zakladki][lek][cegla][2] = round(wart3, 2)
                                if hit_number == 0:
                                    break
