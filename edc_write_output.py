# ! /usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl
from string import ascii_uppercase

__author__ = 'Marcin Pieczyński'


class WriteOutputBase(object):
    def __init__(self):
        self.output_file = None
        self.output_sheets = None

    def create_output_file(self):
        """
        Method for creating output file
        """
        self.output_file = openpyxl.Workbook()

    def save_output(self, output_file_path):
        """
        Method for saving excel output file in a given path.
        """
        self.output_file.save(output_file_path)

    def remowe_first_sheet(self):
        """
        Method remove first sheet in output file with defaults name
        """
        self.output_file.remove_sheet(self.output_file.get_sheet_by_name('Sheet'))
        self.output_sheets = self.output_file.get_sheet_names()


class WriteOutputPnVaUnitsPaternA(WriteOutputBase):
    """
    Class containing methods for writing PnVa and UNITS data to excel output file.
    """
    def __init__(self, output_zakladki, slownik, output_lista_cegiel, output_leki):
        WriteOutputBase.__init__(self)

        self.output_zakladki = output_zakladki
        self.slownik = slownik
        self.output_lista_cegiel = output_lista_cegiel
        self.output_leki = output_leki

        self.alfabet = ascii_uppercase
        self.liczba_pelnych_linii = None
        self.niepelna_liniia = None

    def start_output(self):
        """
        Method for creating output file with sheets.
        """
        self.create_output_file()

        for elem in range(len(self.output_zakladki)):
            self.output_file.create_sheet(self.output_zakladki[elem], elem)

        self.remowe_first_sheet()

    def write_base_data(self, daty):
        """
        Method for writing basa data (dates, cegla and drugs names)
        """
        self.liczba_pelnych_linii = len(self.output_lista_cegiel) / 5
        if len(self.output_lista_cegiel)-self.liczba_pelnych_linii * 5 != 0:
            self.niepelna_liniia = True
        else:
            self.niepelna_liniia = False

        self.write_drugs_name()
        self.write_cegla_name()
        self.write_date(daty)

    def write_drugs_name(self):
        """
        Method for writing basa data - drugs names
        """
        for elem in range(len(self.output_zakladki)):
            n, first_row = 0, 3

            sheet = self.output_zakladki[elem]
            sh = self.output_file.get_sheet_by_name(sheet)
            no_of_row = self.liczba_pelnych_linii

            if self.niepelna_liniia:
                no_of_row += 1

            while no_of_row != 0:
                for lek in self.output_leki[elem]:
                    sh['A' + str(first_row + n)] = lek
                    n += 1
                first_row += 2
                no_of_row -= 1

    def write_cegla_name(self):
        """
        Method for writing basa data - cegla names
        """
        for elem in range(len(self.output_zakladki)):
            first_row, first_col = 1, 3

            sheet = self.output_zakladki[elem]
            sh = self.output_file.get_sheet_by_name(sheet)
            cegla_no = 0

            for no in range(self.liczba_pelnych_linii):         # writing data in complete row, containing 5 cegla
                for i in range(0, 5):
                    sh[self.alfabet[first_col + i] + str(first_row)] = self.output_lista_cegiel[cegla_no]
                    first_col += 2
                    cegla_no += 1
                first_col = 3
                first_row += len(self.output_leki[elem]) + 2

            # writing data in incomplete row, containing < 5 cegla
            for i in range(len(self.output_lista_cegiel) - cegla_no):
                sh[self.alfabet[first_col] + str(first_row)] = self.output_lista_cegiel[cegla_no]
                cegla_no += 1
                first_col += 3

    def write_date(self, daty):
        """
        Method for writing basa data - dates
        """
        for elem in range(len(self.output_zakladki)):
            first_row, first_col, no_date, cegla_no = 2, 1, 0, 0

            sheet = self.output_zakladki[elem]
            sh = self.output_file.get_sheet_by_name(sheet)

            for no in range(self.liczba_pelnych_linii):      # writing data in complete row, containing 5 cegla
                for i in range(15):
                    sh[self.alfabet[first_col] + str(first_row)] = daty[no_date]
                    first_col += 1
                    no_date += 1
                    if no_date == 3:
                        no_date = 0
                cegla_no += 5
                first_col = 1
                first_row += len(self.output_leki[elem]) + 2

            # writing data in incomplete row, containing < 5 cegla
            mising_date = (len(self.output_lista_cegiel) - cegla_no) * 3
            for i in range(mising_date):
                sh[self.alfabet[first_col] + str(first_row)] = daty[no_date]
                first_col += 1
                no_date += 1
                if no_date == 3:
                    no_date = 0

    def write_pnva_units_data(self):
        """
        Method for writing in side output file PnVa and UNITS data.
        """
        for elem in range(len(self.output_zakladki)):
            cegla_no, col, row = 0, 1, 3
            no_of_row = len(self.output_lista_cegiel) / 5

            sheet = self.output_zakladki[elem]
            sh = self.output_file.get_sheet_by_name(sheet)

            for e in range(no_of_row):                    # writing data in complete row, containing 5 cegla
                for lek in self.output_leki[elem]:
                    for i in range(5):
                        for ii in range(3):
                            wart = self.slownik[sheet][lek][self.output_lista_cegiel[i]][ii]
                            sh[self.alfabet[col] + str(row)] = wart
                            col += 1
                        cegla_no += 1
                    col = 1
                    row += 1
                row += 2

            mising_cegla = (len(self.output_lista_cegiel) - (cegla_no / len(self.output_leki[elem])))
            tru_cegla_no = (cegla_no / len(self.output_leki[elem]))

            for lek in self.output_leki[elem]:            # writing data in incomplete row, containing < 5 cegla
                for i in range(mising_cegla):
                    for ii in range(3):
                        wart = self.slownik[sheet][lek][self.output_lista_cegiel[tru_cegla_no]][ii]
                        sh[self.alfabet[col] + str(row)] = wart
                        col += 1
                    tru_cegla_no += 1
                col = 1
                row += 1
                tru_cegla_no = (cegla_no / len(self.output_leki[elem]))


class WriteOutputCeglyPaternA(WriteOutputBase):
    """
    Class containing methods for writing cegla data to excel output file.
    """
    def __init__(self):
        WriteOutputBase.__init__(self)
        self.alfabet = ascii_uppercase

    def start_output(self, output_lista_cegiel):
        """
        Method for creating output file with sheets.
        """
        self.create_output_file()

        for elem in range(len(output_lista_cegiel)):
            self.output_file.create_sheet(output_lista_cegiel[elem], elem)

        self.remowe_first_sheet()

    def write_base_data(self, output_lista_cegiel, output_leki_cegly, daty):
        """
        Methods for writing base data to output file with cegla data.
        """
        for elem in range(len(output_lista_cegiel)):
            sheet = output_lista_cegiel[elem]
            sh = self.output_file.get_sheet_by_name(sheet)
            col = 1
            sh['B3'] = "UNITS"
            sh['E3'] = "PnVa"

            for i in range(2):
                for ii in range(3):
                    sh[self.alfabet[col] + '4'] = daty[ii]
                    col += 1
            col, row = 0, 5
            for i in output_leki_cegly:
                sh[self.alfabet[col] + str(row)] = i
                row += 1

    def write_pnva_units_data(self, output_lista_cegiel, output_leki_cegly, cegly_data):
        """
        Methods for writing PnVa and UNITS data to output file with UNITS data.
        """
        for elem in range(len(output_lista_cegiel)):
            col, row = 1, 5
            sheet = output_lista_cegiel[elem]
            sh = self.output_file.get_sheet_by_name(sheet)

            for i in output_leki_cegly:
                for x in range(3):
                    sh[self.alfabet[col] + str(row)] = cegly_data[sheet][i]["UNITS"][x]
                    col += 1
                for x in range(3):
                    sh[self.alfabet[col] + str(row)] = cegly_data[sheet][i]["PnVa"][x]
                    col += 1
                col = 1
                row += 1
