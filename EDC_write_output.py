# ! /usr/bin/env python
# -*- coding: utf-8 -*-
__author__ = 'Marcin Pieczyński'

import openpyxl
from edc_sqlite import SQliteEdit
from edc_read_input import Read_input

input_file_path1 = "/home/marcin/Pulpit/ExcelDataConverter_2.1/Raporty 3V2015/A DYMOW SZEMBEK SZYMEL Konkurencja Rx 2015_1Q.xlsx"



class Write_output_PnVa_Units_paternA(Read_input):
    """klasa zawierająca metody zapisywania danych w pliku excel output dla danych PnVa oraz UNITS"""

    def __init__(self, rodzaj_danych):
        Read_input.__init__(self, input_file_path1)

        # linijka dodana tylko by zaoszczędzić czas !!!!!!!
        self.PnVa_data = {u'co-amlesa': {'CO-AMLESSA TABS 8 MG /+10 30 +2.5': {u'A4.06': [118.17, 135.85, 40.8], u'A4.07': [None, None, None], u'A4.04': [48.52, 39.5, 0.0], u'A4.05': [0.0, 48.69, 40.39], u'A4.02': [94.72, 54.0, 0.0], u'A4.03': [0.0, 51.45, 70.47], u'A4.01': [0.0, 0.0, 0.0], u'A4.08': [33.17, 0.0, 70.1], u'A4': [37.34, 42.37, 37.67]}, 'CO-AMLESSA TABS 4 MG /+5 30 +1.2': {u'A4.06': [100.29, 0.0, 48.47], u'A4.07': [62.31, 0.0, 47.82], u'A4.04': [0.0, 0.0, 0.0], u'A4.05': [24.58, 0.0, 31.99], u'A4.02': [40.2, 15.37, 28.64], u'A4.03': [0.0, 74.19, 69.55], u'A4.01': [62.67, 113.05, 150.47], u'A4.08': [21.12, 7.5, 48.24], u'A4': [36.45, 17.51, 46.22]}, 'CO-AMLESSA TOTAL': {u'A4.06': [86.87, 43.44, 49.46], u'A4.07': [54.42, -12.29, 22.05], u'A4.04': [23.58, 45.78, 7.78], u'A4.05': [32.84, 28.54, 35.19], u'A4.02': [44.89, 22.66, 17.92], u'A4.03': [40.53, 42.49, 77.0], u'A4.01': [54.73, 39.69, 52.32], u'A4.08': [18.74, 6.85, 51.9], u'A4': [41.34, 24.46, 41.29]}, 'CO-AMLESSA TABS 4 MG /+10 30 +1.2': {u'A4.06': [0.0, 0.0, 171.93], u'A4.07': [410.14, -149.5, 73.34], u'A4.04': [139.9, 96.0, 105.22], u'A4.05': [163.5, 157.8, 147.99], u'A4.02': [0.0, 0.0, 0.0], u'A4.03': [None, None, None], u'A4.01': [412.5, 0.0, 0.0], u'A4.08': [0.0, 0.0, 64.22], u'A4': [104.98, 22.63, 81.83]}, 'CO-AMLESSA TABS 8 MG /+5 30 +2.5': {u'A4.06': [40.87, 0.0, 26.18], u'A4.07': [None, None, None], u'A4.04': [0.0, 102.16, 0.0], u'A4.05': [53.41, 0.0, 0.0], u'A4.02': [0.0, 0.0, 31.45], u'A4.03': [192.26, 0.0, 119.33], u'A4.01': [0.0, 0.0, 0.0], u'A4.08': [0.0, 17.05, 29.99], u'A4': [36.16, 11.78, 27.48]}}, u'roswera': {'SUVARDIO BRAND': {u'A4.06': [134.45, 180.28, 167.92], u'A4.07': [104.1, 108.98, 75.55], u'A4.04': [99.29, 95.22, 145.83], u'A4.05': [142.43, 145.96, 145.78], u'A4.02': [210.66, 169.69, 171.97], u'A4.03': [195.36, 210.2, 206.22], u'A4.01': [198.94, 215.06, 200.03], u'A4.08': [133.08, 129.94, 153.28], u'A4': [144.88, 149.79, 151.97]}, 'ROMAZIC BRAND': {u'A4.06': [61.39, 117.51, 50.14], u'A4.07': [80.61, 71.2, 42.54], u'A4.04': [57.04, 89.79, 77.03], u'A4.05': [69.18, 46.87, 69.31], u'A4.02': [92.43, 74.02, 87.03], u'A4.03': [56.06, 52.76, 53.22], u'A4.01': [73.9, 25.51, 65.94], u'A4.08': [119.74, 109.79, 112.42], u'A4': [81.68, 79.0, 73.46]}, 'ZAHRON BRAND': {u'A4.06': [75.03, 66.5, 74.36], u'A4.07': [93.47, 85.02, 98.7], u'A4.04': [110.41, 63.12, 106.28], u'A4.05': [71.6, 70.5, 79.07], u'A4.02': [73.39, 77.82, 113.67], u'A4.03': [108.88, 118.6, 129.11], u'A4.01': [71.37, 106.25, 129.92], u'A4.08': [79.24, 104.31, 115.31], u'A4': [84.18, 87.0, 102.56]}, 'ZARANTA BRAND': {u'A4.06': [37.29, 42.73, 33.47], u'A4.07': [36.68, 38.08, 25.66], u'A4.04': [36.09, 37.46, 40.51], u'A4.05': [34.69, 25.06, 33.34], u'A4.02': [38.24, 51.73, 55.49], u'A4.03': [62.02, 64.16, 96.97], u'A4.01': [76.53, 47.83, 37.32], u'A4.08': [49.08, 41.67, 50.78], u'A4': [43.86, 41.66, 46.52]}, 'ROSWERA BRAND': {u'A4.06': [153.32, 146.11, 150.42], u'A4.07': [70.02, 69.42, 81.29], u'A4.04': [98.4, 92.3, 121.76], u'A4.05': [100.96, 99.93, 109.05], u'A4.02': [107.7, 138.7, 126.84], u'A4.03': [94.16, 99.65, 97.15], u'A4.01': [108.9, 137.62, 98.92], u'A4.08': [120.18, 118.63, 123.61], u'A4': [107.7, 110.04, 114.31]}}, u'karbis': {'KARBIS TOTAL': {u'A4.06': [39.72, 44.78, 67.81], u'A4.07': [44.2, 41.02, 42.53], u'A4.04': [53.75, 62.35, 46.53], u'A4.05': [56.68, 67.84, 86.51], u'A4.02': [83.7, 73.36, 75.05], u'A4.03': [44.81, 63.72, 78.75], u'A4.01': [140.91, 29.27, 95.62], u'A4.08': [83.98, 103.45, 69.15], u'A4': [65.2, 68.96, 70.47]}, 'LOSARTAN KRKA TOTAL': {u'A4.06': [None, None, None], u'A4.07': [None, None, None], u'A4.04': [None, None, None], u'A4.05': [None, None, None], u'A4.02': [None, None, None], u'A4.03': [None, None, None], u'A4.01': [None, None, None], u'A4.08': [None, None, None], u'A4': [None, None, None]}, 'COZAAR TOTAL': {u'A4.06': [None, None, None], u'A4.07': [None, None, None], u'A4.04': [None, None, None], u'A4.05': [None, None, None], u'A4.02': [None, None, None], u'A4.03': [None, None, None], u'A4.01': [None, None, None], u'A4.08': [None, None, None], u'A4': [None, None, None]}, 'LOSARTAN KRKA FILM C. TABS 50 MG 30': {u'A4.06': [None, None, None], u'A4.07': [None, None, None], u'A4.04': [None, None, None], u'A4.05': [None, None, None], u'A4.02': [None, None, None], u'A4.03': [None, None, None], u'A4.01': [None, None, None], u'A4.08': [None, None, None], u'A4': [None, None, None]}, 'CARZAP BRAND': {u'A4.06': [17.97, 47.96, 50.57], u'A4.07': [0.0, 7.58, 6.04], u'A4.04': [82.54, 75.41, 86.03], u'A4.05': [33.52, 21.42, 5.53], u'A4.02': [58.05, 0.0, 19.01], u'A4.03': [128.06, 212.83, 83.24], u'A4.01': [178.99, 20.04, 37.48], u'A4.08': [6.91, 19.99, 16.26], u'A4': [45.56, 46.32, 32.01]}, 'LORISTA TOTAL': {u'A4.06': [None, None, None], u'A4.07': [None, None, None], u'A4.04': [None, None, None], u'A4.05': [None, None, None], u'A4.02': [None, None, None], u'A4.03': [None, None, None], u'A4.01': [None, None, None], u'A4.08': [None, None, None], u'A4': [None, None, None]}, 'TOLURA TOTAL': {u'A4.06': [57.07, 69.38, 97.32], u'A4.07': [59.25, 58.33, 81.69], u'A4.04': [68.52, 24.12, 64.49], u'A4.05': [64.95, 73.59, 62.45], u'A4.02': [107.0, 102.93, 95.39], u'A4.03': [52.48, 53.12, 74.65], u'A4.01': [28.44, 43.73, 24.22], u'A4.08': [78.97, 71.3, 72.45], u'A4': [68.16, 66.88, 74.66]}}, u'sobycombi': {'SOBYCOMBI TABS 10MG /+5 30': {u'A4.06': [0.0, 0.0, 327.08], u'A4.07': [277.5, -308.53, 0.0], u'A4.04': [0.0, 0.0, 0.0], u'A4.05': [0.0, 112.15, 0.0], u'A4.02': [None, None, None], u'A4.03': [0.0, 0.0, 0.0], u'A4.01': [0.0, 0.0, 0.0], u'A4.08': [0.0, 0.0, 66.01], u'A4': [26.85, 1.53, 56.17]}, 'SOBYCOMBI TABS 10MG /+10 30': {u'A4.06': [0.0, 286.67, 132.03], u'A4.07': [266.25, -239.76, 0.0], u'A4.04': [0.0, 0.0, 0.0], u'A4.05': [0.0, 85.42, 0.0], u'A4.02': [0.0, 0.0, 301.53], u'A4.03': [0.0, 0.0, 115.87], u'A4.01': [0.0, 0.0, 0.0], u'A4.08': [0.0, 90.05, 0.0], u'A4': [25.76, 56.67, 60.46]}, 'SOBYCOMBI TOTAL': {u'A4.06': [-136.09, 95.84, 19.37], u'A4.07': [153.92, -102.02, 67.41], u'A4.04': [0.0, 34.19, 0.0], u'A4.05': [-43.05, 50.91, 39.66], u'A4.02': [-161.48, 25.97, 126.73], u'A4.03': [-52.29, 21.14, 45.11], u'A4.01': [0.0, -27.76, 0.0], u'A4.08': [49.09, -0.65, 70.96], u'A4': [-12.5, 22.47, 53.89]}, 'SOBYCOMBI TABS 5MG /+10 30': {u'A4.06': [-1613.84, 475.08, -85.36], u'A4.07': [820.58, -267.05, -129.83], u'A4.04': [0.0, 278.4, 0.0], u'A4.05': [-733.76, 96.06, 57.07], u'A4.02': [-4332.06, -223.94, 326.83], u'A4.03': [-1527.91, 0.0, 0.0], u'A4.01': [0.0, 0.0, 0.0], u'A4.08': [15.72, -383.77, 113.5], u'A4': [-743.84, -15.43, 55.76]}, 'SOBYCOMBI TABS 5MG /+5 30': {u'A4.06': [-91.33, -18.36, -12.61], u'A4.07': [60.49, 0.0, 135.05], u'A4.04': [0.0, 0.0, 0.0], u'A4.05': [-14.27, 23.87, 48.69], u'A4.02': [42.14, 85.68, 60.22], u'A4.03': [25.22, 33.87, 46.29], u'A4.01': [0.0, -44.49, 0.0], u'A4.08': [72.04, 51.39, 75.43], u'A4': [20.22, 23.85, 52.11]}}, u'atoris': {'ATORVASTEROL TOTAL': {u'A4.06': [105.11, 88.6, 89.31], u'A4.07': [72.78, 62.62, 85.01], u'A4.04': [65.67, 45.57, 70.87], u'A4.05': [72.6, 70.17, 67.04], u'A4.02': [69.48, 41.71, 46.65], u'A4.03': [65.58, 57.08, 76.96], u'A4.01': [65.18, 63.54, 47.02], u'A4.08': [94.12, 92.18, 67.28], u'A4': [79.94, 70.29, 71.15]}, 'TORVACARD TOTAL': {u'A4.06': [None, None, None], u'A4.07': [None, None, None], u'A4.04': [None, None, None], u'A4.05': [None, None, None], u'A4.02': [None, None, None], u'A4.03': [None, None, None], u'A4.01': [None, None, None], u'A4.08': [None, None, None], u'A4': [None, None, None]}, 'TORVALIPIN BRAND': {u'A4.06': [None, None, None], u'A4.07': [None, None, None], u'A4.04': [None, None, None], u'A4.05': [None, None, None], u'A4.02': [None, None, None], u'A4.03': [None, None, None], u'A4.01': [None, None, None], u'A4.08': [None, None, None], u'A4': [None, None, None]}, 'ATORIS FILM C.TABS 30 MG 30': {u'A4.06': [60.02, 35.77, 45.12], u'A4.07': [51.72, 129.69, 11.95], u'A4.04': [62.42, 25.66, 94.85], u'A4.05': [25.51, 34.66, 22.89], u'A4.02': [11.65, 56.54, 15.32], u'A4.03': [34.29, 34.47, 30.04], u'A4.01': [46.53, 98.0, 94.47], u'A4.08': [56.55, 43.63, 50.21], u'A4': [44.22, 53.97, 38.45]}, 'ATORIS FILM C.TABS 30 MG 60': {u'A4.06': [14.03, 7.88, 13.35], u'A4.07': [62.73, 74.95, 58.04], u'A4.04': [47.13, 13.21, 78.69], u'A4.05': [19.34, 51.18, 46.73], u'A4.02': [19.83, 72.13, -62.66], u'A4.03': [23.21, 41.27, 27.33], u'A4.01': [45.15, 0.0, 64.46], u'A4.08': [33.23, 57.44, 44.91], u'A4': [31.61, 46.83, 34.71]}, 'TULIP TOTAL': {u'A4.06': [130.5, 131.07, 162.87], u'A4.07': [141.72, 153.79, 140.25], u'A4.04': [114.96, 128.17, 163.6], u'A4.05': [122.01, 128.06, 140.07], u'A4.02': [133.58, 134.4, 126.95], u'A4.03': [100.57, 117.41, 100.57], u'A4.01': [130.43, 103.17, 108.91], u'A4.08': [114.98, 112.32, 119.8], u'A4': [122.55, 126.79, 132.8]}, 'ATORIS TOTAL': {u'A4.06': [112.16, 107.34, 107.78], u'A4.07': [80.81, 88.55, 80.69], u'A4.04': [96.97, 101.97, 90.51], u'A4.05': [101.6, 103.31, 93.43], u'A4.02': [106.75, 117.04, 116.23], u'A4.03': [89.07, 91.37, 88.49], u'A4.01': [102.07, 72.37, 96.5], u'A4.08': [95.55, 96.22, 91.68], u'A4': [97.28, 98.7, 94.35]}}, u'aclexa': {'ACLEXA CAPS 100MG 10': {u'A4.06': [103.1, 195.21, 112.11], u'A4.07': [0.0, -52.37, 1.5], u'A4.04': [35.13, 459.32, 146.75], u'A4.05': [-10.97, 5.13, 73.05], u'A4.02': [0.0, 88.87, 396.15], u'A4.03': [41.05, 31.02, 110.77], u'A4.01': [-58.65, 0.0, 851.81], u'A4.08': [10.36, 40.84, 120.4], u'A4': [18.91, 67.58, 144.03]}, 'ACLEXA TOTAL': {u'A4.06': [34.39, 23.14, 64.55], u'A4.07': [0.0, 107.45, 89.3], u'A4.04': [2.68, 66.15, 53.22], u'A4.05': [6.58, 16.95, 48.04], u'A4.02': [32.83, 59.54, 93.82], u'A4.03': [7.05, 20.41, 57.52], u'A4.01': [8.32, 13.99, 125.82], u'A4.08': [11.18, 10.0, 31.08], u'A4': [12.85, 37.34, 60.46]}, 'NALGESIN TOTAL': {u'A4.06': [60.23, 85.82, 52.69], u'A4.07': [18.59, 20.19, 28.32], u'A4.04': [61.36, 54.47, 62.69], u'A4.05': [85.15, 64.96, 64.62], u'A4.02': [52.0, 100.18, 60.9], u'A4.03': [45.5, 38.37, 45.24], u'A4.01': [80.74, 85.92, 107.3], u'A4.08': [52.62, 112.22, 80.61], u'A4': [54.83, 71.86, 60.39]}, 'ANAPRAN TOTAL': {u'A4.06': [63.05, 76.71, 73.44], u'A4.07': [188.18, 196.37, 182.08], u'A4.04': [88.5, 134.3, 92.86], u'A4.05': [81.64, 60.19, 50.03], u'A4.02': [116.43, 152.36, 104.14], u'A4.03': [55.71, 62.79, 80.19], u'A4.01': [85.95, 83.07, 72.74], u'A4.08': [72.72, 113.9, 72.8], u'A4': [95.22, 109.18, 89.49]}, 'NAKLOFEN TOTAL': {u'A4.06': [9.41, 23.28, 36.44], u'A4.07': [23.63, 40.6, 26.19], u'A4.04': [11.0, 9.31, 24.42], u'A4.05': [24.02, 15.03, 35.29], u'A4.02': [15.34, 28.77, 21.67], u'A4.03': [20.68, 10.42, 4.35], u'A4.01': [7.41, 8.8, 6.13], u'A4.08': [22.78, 35.31, 23.54], u'A4': [19.07, 24.65, 24.99]}, 'APO-NAPRO TOTAL': {u'A4.06': [173.36, 205.79, 165.48], u'A4.07': [83.39, 86.02, 77.26], u'A4.04': [174.54, 194.56, 133.16], u'A4.05': [180.38, 169.19, 138.36], u'A4.02': [102.03, 125.56, 91.85], u'A4.03': [107.56, 149.95, 141.21], u'A4.01': [168.36, 192.28, 175.51], u'A4.08': [182.88, 229.9, 142.71], u'A4': [147.92, 170.48, 130.72]}, 'ACLEXA CAPS 200MG 10': {u'A4.06': [129.17, 114.66, 178.78], u'A4.07': [0.0, 292.54, 119.48], u'A4.04': [0.0, 87.56, 56.86], u'A4.05': [24.94, -30.64, 49.86], u'A4.02': [20.75, 14.34, 83.23], u'A4.03': [0.0, 13.01, 39.27], u'A4.01': [0.0, -35.56, 28.62], u'A4.08': [0.0, -4.19, 33.47], u'A4': [23.96, 58.69, 74.21]}, 'ACLEXA CAPS 100MG 60': {u'A4.06': [25.98, 0.0, 29.19], u'A4.07': [0.0, 0.0, 50.66], u'A4.04': [0.0, 119.88, 30.29], u'A4.05': [0.0, 45.75, 77.27], u'A4.02': [33.81, 109.73, 58.46], u'A4.03': [0.0, 0.0, 16.74], u'A4.01': [0.0, 55.45, 221.77], u'A4.08': [47.98, 0.0, 23.51], u'A4': [17.43, 29.22, 48.08]}, 'ACLEXA CAPS 200MG 60': {u'A4.06': [0.0, 0.0, 25.45], u'A4.07': [0.0, 0.0, 22.09], u'A4.04': [0.0, 0.0, 52.82], u'A4.05': [0.0, 37.19, 33.69], u'A4.02': [0.0, 74.33, 84.97], u'A4.03': [0.0, 67.46, 102.16], u'A4.01': [0.0, 0.0, 0.0], u'A4.08': [0.0, 0.0, 13.67], u'A4': [0.0, 22.27, 38.7]}, 'ACLEXA CAPS 100MG 30': {u'A4.06': [55.44, 25.16, 54.78], u'A4.07': [0.0, 32.05, -14.86], u'A4.04': [0.0, 25.44, 94.14], u'A4.05': [0.0, 40.38, 143.88], u'A4.02': [0.0, 96.24, 156.59], u'A4.03': [32.19, 73.26, 83.43], u'A4.01': [0.0, 98.83, 276.39], u'A4.08': [33.32, 52.07, 39.52], u'A4': [18.24, 49.91, 80.69]}, 'ACLEXA CAPS 200MG 30': {u'A4.06': [8.27, 0.0, 63.52], u'A4.07': [0.0, 173.61, 176.93], u'A4.04': [0.0, 34.72, 38.05], u'A4.05': [12.32, 7.07, 21.09], u'A4.02': [64.59, 35.31, 60.34], u'A4.03': [0.0, 0.0, 31.29], u'A4.01': [32.94, 0.0, 86.24], u'A4.08': [0.0, 9.74, 31.65], u'A4': [11.1, 35.2, 59.77]}}, u'karbikombi': {'KARBICOMBI TOTAL': {u'A4.06': [40.08, 40.87, 25.53], u'A4.07': [70.42, 50.05, 59.13], u'A4.04': [34.31, 41.07, 53.13], u'A4.05': [73.29, 93.99, 70.13], u'A4.02': [67.74, 84.04, 42.57], u'A4.03': [33.03, 39.42, 29.6], u'A4.01': [135.11, 105.07, 84.11], u'A4.08': [110.17, 70.57, 113.0], u'A4': [70.82, 64.61, 61.49]}, 'ALORTIA TABS 100 MG /+5 30': {u'A4.06': [None, None, None], u'A4.07': [None, None, None], u'A4.04': [None, None, None], u'A4.05': [0.0, 0.0, 0.0], u'A4.02': [None, None, None], u'A4.03': [None, None, None], u'A4.01': [None, None, None], u'A4.08': [None, None, None], u'A4': [0.0, 0.0, 0.0]}, 'ALORTIA TOTAL': {u'A4.06': [-37.3, 0.0, 0.0], u'A4.07': [0.0, 10.49, 22.58], u'A4.04': [0.0, 0.0, 0.0], u'A4.05': [-100.9, 0.0, 19.58], u'A4.02': [-161.42, 0.0, -33.2], u'A4.03': [0.0, 18.58, -20.74], u'A4.01': [0.0, 0.0, 41.66], u'A4.08': [0.0, 18.39, -37.46], u'A4': [-40.25, 7.35, -6.63]}, 'TWYNSTA BRAND': {u'A4.06': [207.05, 193.6, 269.02], u'A4.07': [13.84, 9.85, 4.48], u'A4.04': [135.73, 21.21, 63.06], u'A4.05': [41.44, 35.68, -13.58], u'A4.02': [28.04, 30.7, 18.75], u'A4.03': [144.0, 306.19, 344.03], u'A4.01': [27.75, 0.0, 0.0], u'A4.08': [6.78, 11.02, 11.18], u'A4': [73.56, 87.86, 98.22]}, 'ALORTIA TABS 100 MG /+10 30': {u'A4.06': [None, None, None], u'A4.07': [0.0, 0.0, 309.18], u'A4.04': [None, None, None], u'A4.05': [None, None, None], u'A4.02': [None, None, None], u'A4.03': [None, None, None], u'A4.01': [None, None, None], u'A4.08': [0.0, 0.0, 0.0], u'A4': [0.0, 0.0, 32.54]}, 'EXFORGE BRAND': {u'A4.06': [134.81, 169.96, 235.9], u'A4.07': [64.6, 51.16, 46.44], u'A4.04': [61.15, 46.56, 57.45], u'A4.05': [150.13, 108.52, 123.84], u'A4.02': [77.85, 69.3, 130.75], u'A4.03': [140.65, 118.9, 92.55], u'A4.01': [0.0, 85.41, 33.83], u'A4.08': [170.38, 127.1, 203.51], u'A4': [118.93, 108.13, 137.72]}, 'ALORTIA TABS 50 MG /+10 30': {u'A4.06': [-211.26, 0.0, 0.0], u'A4.07': [0.0, 0.0, -145.3], u'A4.04': [0.0, 0.0, 0.0], u'A4.05': [-576.63, 0.0, 101.3], u'A4.02': [-914.14, 0.0, -156.1], u'A4.03': [0.0, 66.55, 7.88], u'A4.01': [0.0, 0.0, 0.0], u'A4.08': [0.0, 47.43, -142.45], u'A4': [-228.81, 18.88, -41.51]}, 'LOZAP HCT TOTAL': {u'A4.06': [50.91, 51.25, 58.47], u'A4.07': [53.28, 34.69, 37.02], u'A4.04': [50.59, 45.49, 50.32], u'A4.05': [33.97, 28.41, 34.74], u'A4.02': [22.03, 39.32, 39.77], u'A4.03': [64.34, 55.86, 63.68], u'A4.01': [24.29, 34.92, 93.73], u'A4.08': [82.45, 72.76, 61.48], u'A4': [51.99, 48.12, 52.23]}, 'LORISTA H TOTAL': {u'A4.06': [88.06, 114.66, 64.21], u'A4.07': [89.05, 51.43, 82.4], u'A4.04': [119.31, 88.84, 128.87], u'A4.05': [113.33, 107.29, 87.38], u'A4.02': [93.07, 101.29, 81.26], u'A4.03': [100.57, 66.55, 84.99], u'A4.01': [92.07, 37.81, 84.59], u'A4.08': [77.25, 106.28, 76.41], u'A4': [94.78, 91.98, 83.27]}, 'ALORTIA TABS 50 MG /+5 30': {u'A4.06': [-24.16, 0.0, 0.0], u'A4.07': [0.0, 21.22, -24.22], u'A4.04': [0.0, 0.0, 0.0], u'A4.05': [-65.95, 0.0, 16.12], u'A4.02': [-104.55, 0.0, -26.02], u'A4.03': [0.0, 0.0, -36.76], u'A4.01': [0.0, 0.0, 77.19], u'A4.08': [0.0, 11.16, -36.94], u'A4': [-26.17, 4.35, -12.28]}, 'TOLUCOMBI TOTAL': {u'A4.06': [28.51, 35.79, 33.24], u'A4.07': [32.07, 32.69, 35.2], u'A4.04': [69.26, 32.87, 58.02], u'A4.05': [37.88, 63.01, 60.04], u'A4.02': [86.51, 87.01, 101.42], u'A4.03': [17.76, 23.92, 36.88], u'A4.01': [21.93, 13.74, 0.0], u'A4.08': [74.7, 86.09, 81.54], u'A4': [48.15, 53.28, 56.1]}}, u'gliclada': {'GLICLADA TOTAL': {u'A4.06': [74.79, 92.69, 101.69], u'A4.07': [33.75, 28.37, 28.66], u'A4.04': [65.49, 110.29, 130.58], u'A4.05': [68.56, 108.69, 90.44], u'A4.02': [69.91, 98.15, 96.35], u'A4.03': [87.25, 158.62, 103.11], u'A4.01': [50.67, 82.99, 77.75], u'A4.08': [78.31, 82.54, 60.33], u'A4': [66.79, 93.53, 80.13]}, 'DIAPREL MR TOTAL': {u'A4.06': [111.69, 119.45, 116.91], u'A4.07': [107.17, 118.62, 109.5], u'A4.04': [113.26, 112.49, 113.37], u'A4.05': [104.9, 108.64, 108.1], u'A4.02': [127.78, 123.33, 134.87], u'A4.03': [109.61, 108.27, 121.29], u'A4.01': [126.72, 101.37, 128.28], u'A4.08': [109.3, 107.59, 117.11], u'A4': [111.56, 112.3, 116.49]}}}
        # linijka dodana tylko by zaoszczędzić czas !!!!!!!

        # BLOK TESTOWY !!!!!!!!!!!!!!!!!!!!!
        # self.output_lista_cegiel = [u'A4', u'A4.01', u'A4.02', u'A4.03', u'A4.04', u'A4.05', u'A4.06', u'A4.07', u'A4.08',
        #                             u'A4', u'A4.01', u'A4.02', u'A4.03', u'A4.04', u'A4.05', u'A4.06', u'A4.07', u'A4.08']
        # #
        # self.output_lista_cegiel = [u'A4', u'A4.01', u'A4.02']
        # #
        # self.output_lista_cegiel = [u'A4', u'A4.01', u'A4.02', u'A4.03', u'A4.04',
        #                             u'A4.05', u'A4.06', u'A4.07', u'A4.08', u'A4']
        # BLOK TESTOWY !!!!!!!!!!!!!!!!!!!!!
        # print "output_lista_cegiel", self.output_lista_cegiel
        # print self.output_lista_cegiel[5]

        if rodzaj_danych == "PnVa":
            self.slownik = self.PnVa_data
        elif rodzaj_danych == "UNITS":
            self.slownik = self.UNITS_data
        # print self.slownik

        self.output_file = None

        # self.start_output()
        # self.write_base_data()
        # self.write_PnVa_UNITS_data()
        # self.save_output()


    def start_output(self):
        """inicjalizacja pliku output i tworzenie w nim zakłądek"""

        self.output_file = openpyxl.Workbook()
        for elem in range(len(self.output_zakladki)):
            self.output_file.create_sheet(elem, self.output_zakladki[elem])

        # usuwa pierwszą z stworzonych domyślną zakładkę - pewnie można to jakoś obejść .....
        self.output_file.remove_sheet(self.output_file.get_sheet_by_name('Sheet'))
        self.output_sheets = self.output_file.get_sheet_names()


    def save_output(self):
        """zapisuje dane dla pliku output w podanej lokalizacji"""

        self.output_file.save('test_excel_file.xlsx')


    def write_base_data(self):
        """Zapisuje w zakladkach podstawowe dane (daty, cegły i nazwy leków)"""

        liczba_pelnych_linii = len(self.output_lista_cegiel) / 5
        if len(self.output_lista_cegiel)-liczba_pelnych_linii * 5 != 0:
            niepelna_liniia = True
        else:
            niepelna_liniia = False

        # print "\nliczba_pelnych_linii", liczba_pelnych_linii
        # print "niepelna_liniia", niepelna_liniia

        # zapisywanie danych leków w pliku output w poszczególnych zakładkach
        for elem in range(len(self.output_zakladki)):
            n, first_row = 0, 3

            sheet = self.output_zakladki[elem]
            sh = self.output_file.get_sheet_by_name(sheet)

            no_of_row = liczba_pelnych_linii
            if niepelna_liniia:
                no_of_row += 1

            while no_of_row != 0:
                for lek in self.output_leki[elem]:
                    sh['A' + str(first_row + n)] = lek
                    n += 1
                first_row += 2
                no_of_row -= 1

        # zapisywanie nazw cegiel w pliku output w poszczególnych zakładkach
        for elem in range(len(self.output_zakladki)):
            first_row, first_col = 1, 3

            sheet = self.output_zakladki[elem]
            sh = self.output_file.get_sheet_by_name(sheet)

            cegla_no = 0
            for no in range(liczba_pelnych_linii):                     # drukuje pełne linie liczące po 5 cegieł
                for i in range(0, 5):
                    sh[self.alfabet[first_col + i] + str(first_row)] = self.output_lista_cegiel[cegla_no]
                    first_col += 2
                    cegla_no += 1
                first_col = 3
                first_row += len(self.output_leki[elem]) + 2

            # print len(self.output_lista_cegiel), cegla_no
            for i in range(len(self.output_lista_cegiel) - cegla_no):                 # drukuje NIEpełne linie
                sh[self.alfabet[first_col] + str(first_row)] = self.output_lista_cegiel[cegla_no]
                cegla_no += 1
                first_col += 3

        # zapisywanie dat w pliku output w poszczególnych zakładkach
        for elem in range(len(self.output_zakladki)):
            first_row, first_col = 2, 1

            sheet = self.output_zakladki[elem]
            sh = self.output_file.get_sheet_by_name(sheet)
            no_date, cegla_no = 0, 0

            for no in range(liczba_pelnych_linii):                         # drukuje pełne linie liczące po 5 cegieł
                for i in range(15):
                    sh[self.alfabet[first_col] + str(first_row)] = self.daty[no_date]
                    first_col += 1
                    no_date += 1
                    if no_date == 3: no_date = 0
                cegla_no += 5
                first_col = 1
                first_row += len(self.output_leki[elem]) + 2

            # print len(self.output_lista_cegiel), cegla_no
            mising_date = (len(self.output_lista_cegiel) - cegla_no) * 3
            for i in range(mising_date):                                     # drukuje NIEpełne linie
                sh[self.alfabet[first_col] + str(first_row)] = self.daty[no_date]
                first_col += 1
                no_date += 1
                if no_date == 3: no_date = 0


    def write_PnVa_UNITS_data(self):
        """ Zapisuje w pliku output dane cyfrowe dla PnVa lub UNITS"""

        liczba_pelnych_linii = len(self.output_lista_cegiel) / 5
        if len(self.output_lista_cegiel)-liczba_pelnych_linii * 5 != 0:
            niepelna_liniia = True
        else:
            niepelna_liniia = False

        # print "\nliczba_pelnych_linii", liczba_pelnych_linii
        # print "niepelna_liniia", niepelna_liniia

        for elem in range(len(self.output_zakladki)):
            cegla_no, col, row = 0, 1, 3
            no_of_row = liczba_pelnych_linii

            sheet = self.output_zakladki[elem]
            sh = self.output_file.get_sheet_by_name(sheet)

            for e in range(no_of_row):                   # zapisywanie danych w pełnych liniach, zawierających 5 cegieł
                for lek in self.output_leki[elem]:
                    for i in range(5):
                        for ii in range(3):
                            wart = self.slownik[sheet][lek][self.output_lista_cegiel[i]][ii]
                            # if wart == None:
                            #     wart = "b.d."
                            sh[self.alfabet[col] + str(row)] = wart
                            col += 1
                        cegla_no += 1
                    col = 1
                    row += 1
                row += 2

            mising_cegla = (len(self.output_lista_cegiel) - (cegla_no / len(self.output_leki[elem])))
            tru_cegla_no = (cegla_no / len(self.output_leki[elem]))

            # print len(self.output_lista_cegiel), cegla_no
            # print "row", row, "cegla_no", cegla_no, cegla_no / len(self.output_leki[elem])
            # print "mising_cegla", mising_cegla
            # print "tru_cegla_no", tru_cegla_no
            # print sheet

            for lek in self.output_leki[elem]:             # drukuje NIEpełne linii
                for i in range(mising_cegla):
                    for ii in range(3):
                        print "tru_cegla_no", tru_cegla_no
                        wart = self.slownik[sheet][lek][self.output_lista_cegiel[tru_cegla_no]][ii]
                        sh[self.alfabet[col] + str(row)] = wart
                        col += 1
                    tru_cegla_no += 1
                col = 1
                row += 1
                tru_cegla_no = (cegla_no / len(self.output_leki[elem]))






        # self.PnVa_data = {nazwa grupy leków1: {lek1: {cegła1: wartość, cegła2: wartość, cegła3: wartość,
        #                                              cegła4: wartość, cegła5: wartość, cegła5: wartość,}
        #                                      {lek2: {cegła1: wartość, cegła2: wartość, cegła3: wartość,
        #                                              cegła4: wartość, cegła5: wartość, cegła5: wartość,}
        #                 {nazwa grupy leków2: {lek1: {cegła1: wartość, cegła2: wartość, cegła3: wartość,
        #                                              cegła4: wartość, cegła5: wartość, cegła5: wartość,}
        #                                      {lek2: {cegła1: wartość, cegła2: wartość, cegła3: wartość,
        #                                              cegła4: wartość, cegła5: wartość, cegła5: wartość,}}}


class Write_output_CEGLY_paternA(Read_input):
    """klasa zawierająca metody zapisywania danych w pliku excel output dla danych cegly"""

    def __init__(self):
        Read_input.__init__(self, input_file_path1)
        print self.output_zakladki

        self.CEGLY_data = {u'A4.06': {u'ALORTIA TABS 100 MG /+5 30': {'UNITS': [None, None, None], 'PnVa': [None, None, None]}, u'ACLEXA TOTAL': {'UNITS': [20.0, 22.0, 58.0], 'PnVa': [34.39, 23.14, 64.55]}, u'SOBYCOMBI TABS 5MG /+10 30': {'UNITS': [-2.0, 3.0, -1.0], 'PnVa': [-1613.84, 475.08, -85.36]}, u'SOBYCOMBI TABS 5MG /+5 30': {'UNITS': [-3.0, -1.0, -1.0], 'PnVa': [-91.33, -18.36, -12.61]}, u'KARBIS TOTAL': {'UNITS': [27.0, 27.0, 38.0], 'PnVa': [39.72, 44.78, 67.81]}, u'KARBICOMBI TOTAL': {'UNITS': [23.0, 24.0, 14.0], 'PnVa': [40.08, 40.87, 25.53]}, u'GLICLADA TOTAL': {'UNITS': [58.0, 65.0, 79.0], 'PnVa': [74.79, 92.69, 101.69]}, u'SOBYCOMBI TABS 10MG /+5 30': {'UNITS': [0.0, 0.0, 2.0], 'PnVa': [0.0, 0.0, 327.08]}, u'ATORIS FILM C.TABS 30 MG 30': {'UNITS': [29.0, 14.0, 20.0], 'PnVa': [60.02, 35.77, 45.12]}, u'ROSWERA BRAND': {'UNITS': [939.0, 731.0, 842.0], 'PnVa': [153.32, 146.11, 150.42]}, u'ALORTIA TABS 50 MG /+10 30': {'UNITS': [-1.0, 0.0, 0.0], 'PnVa': [-211.26, 0.0, 0.0]}, u'ALORTIA TABS 100 MG /+10 30': {'UNITS': [None, None, None], 'PnVa': [None, None, None]}, u'SOBYCOMBI TOTAL': {'UNITS': [-5.0, 4.0, 1.0], 'PnVa': [-136.09, 95.84, 19.37]}, u'ATORIS FILM C.TABS 30 MG 60': {'UNITS': [2.0, 1.0, 2.0], 'PnVa': [14.03, 7.88, 13.35]}, u'CO-AMLESSA TOTAL': {'UNITS': [13.0, 6.0, 10.0], 'PnVa': [86.87, 43.44, 49.46]}, u'ALORTIA TOTAL': {'UNITS': [-2.0, 0.0, 0.0], 'PnVa': [-37.3, 0.0, 0.0]}, u'SOBYCOMBI TABS 10MG /+10 30': {'UNITS': [0.0, 2.0, 1.0], 'PnVa': [0.0, 286.67, 132.03]}, u'ALORTIA TABS 50 MG /+5 30': {'UNITS': [-1.0, 0.0, 0.0], 'PnVa': [-24.16, 0.0, 0.0]}, u'ATORIS TOTAL': {'UNITS': [1001.0, 745.0, 903.0], 'PnVa': [112.16, 107.34, 107.78]}}, u'A4.07': {u'ALORTIA TABS 100 MG /+5 30': {'UNITS': [None, None, None], 'PnVa': [None, None, None]}, u'ACLEXA TOTAL': {'UNITS': [0.0, 70.0, 69.0], 'PnVa': [0.0, 107.45, 89.3]}, u'SOBYCOMBI TABS 5MG /+10 30': {'UNITS': [1.0, -1.0, -1.0], 'PnVa': [820.58, -267.05, -129.83]}, u'SOBYCOMBI TABS 5MG /+5 30': {'UNITS': [2.0, 0.0, 7.0], 'PnVa': [60.49, 0.0, 135.05]}, u'KARBIS TOTAL': {'UNITS': [30.0, 23.0, 28.0], 'PnVa': [44.2, 41.02, 42.53]}, u'KARBICOMBI TOTAL': {'UNITS': [28.0, 17.0, 23.0], 'PnVa': [70.42, 50.05, 59.13]}, u'GLICLADA TOTAL': {'UNITS': [41.0, 28.0, 33.0], 'PnVa': [33.75, 28.37, 28.66]}, u'SOBYCOMBI TABS 10MG /+5 30': {'UNITS': [1.0, -1.0, 0.0], 'PnVa': [277.5, -308.53, 0.0]}, u'ATORIS FILM C.TABS 30 MG 30': {'UNITS': [28.0, 59.0, 6.0], 'PnVa': [51.72, 129.69, 11.95]}, u'ROSWERA BRAND': {'UNITS': [433.0, 378.0, 470.0], 'PnVa': [70.02, 69.42, 81.29]}, u'ALORTIA TABS 50 MG /+10 30': {'UNITS': [0.0, 0.0, -1.0], 'PnVa': [0.0, 0.0, -145.3]}, u'ALORTIA TABS 100 MG /+10 30': {'UNITS': [0.0, 0.0, 2.0], 'PnVa': [0.0, 0.0, 309.18]}, u'SOBYCOMBI TOTAL': {'UNITS': [5.0, -3.0, 6.0], 'PnVa': [153.92, -102.02, 67.41]}, u'ATORIS FILM C.TABS 30 MG 60': {'UNITS': [10.0, 11.0, 10.0], 'PnVa': [62.73, 74.95, 58.04]}, u'CO-AMLESSA TOTAL': {'UNITS': [8.0, -2.0, 4.0], 'PnVa': [54.42, -12.29, 22.05]}, u'ALORTIA TOTAL': {'UNITS': [0.0, 1.0, 0.0], 'PnVa': [0.0, 10.49, 22.58]}, u'SOBYCOMBI TABS 10MG /+10 30': {'UNITS': [1.0, -1.0, 0.0], 'PnVa': [266.25, -239.76, 0.0]}, u'ALORTIA TABS 50 MG /+5 30': {'UNITS': [0.0, 1.0, -1.0], 'PnVa': [0.0, 21.22, -24.22]}, u'ATORIS TOTAL': {'UNITS': [726.0, 637.0, 741.0], 'PnVa': [80.81, 88.55, 80.69]}}, u'A4.04': {u'ALORTIA TABS 100 MG /+5 30': {'UNITS': [None, None, None], 'PnVa': [None, None, None]}, u'ACLEXA TOTAL': {'UNITS': [1.0, 22.0, 21.0], 'PnVa': [2.68, 66.15, 53.22]}, u'SOBYCOMBI TABS 5MG /+10 30': {'UNITS': [0.0, 1.0, 0.0], 'PnVa': [0.0, 278.4, 0.0]}, u'SOBYCOMBI TABS 5MG /+5 30': {'UNITS': [0.0, 0.0, 0.0], 'PnVa': [0.0, 0.0, 0.0]}, u'KARBIS TOTAL': {'UNITS': [25.0, 20.0, 18.0], 'PnVa': [53.75, 62.35, 46.53]}, u'KARBICOMBI TOTAL': {'UNITS': [8.0, 12.0, 16.0], 'PnVa': [34.31, 41.07, 53.13]}, u'GLICLADA TOTAL': {'UNITS': [40.0, 56.0, 63.0], 'PnVa': [65.49, 110.29, 130.58]}, u'SOBYCOMBI TABS 10MG /+5 30': {'UNITS': [0.0, 0.0, 0.0], 'PnVa': [0.0, 0.0, 0.0]}, u'ATORIS FILM C.TABS 30 MG 30': {'UNITS': [18.0, 6.0, 25.0], 'PnVa': [62.42, 25.66, 94.85]}, u'ROSWERA BRAND': {'UNITS': [369.0, 300.0, 448.0], 'PnVa': [98.4, 92.3, 121.76]}, u'ALORTIA TABS 50 MG /+10 30': {'UNITS': [0.0, 0.0, 0.0], 'PnVa': [0.0, 0.0, 0.0]}, u'ALORTIA TABS 100 MG /+10 30': {'UNITS': [None, None, None], 'PnVa': [None, None, None]}, u'SOBYCOMBI TOTAL': {'UNITS': [0.0, 1.0, 0.0], 'PnVa': [0.0, 34.19, 0.0]}, u'ATORIS FILM C.TABS 30 MG 60': {'UNITS': [4.0, 1.0, 7.0], 'PnVa': [47.13, 13.21, 78.69]}, u'CO-AMLESSA TOTAL': {'UNITS': [2.0, 4.0, 1.0], 'PnVa': [23.58, 45.78, 7.78]}, u'ALORTIA TOTAL': {'UNITS': [0.0, 0.0, 0.0], 'PnVa': [0.0, 0.0, 0.0]}, u'SOBYCOMBI TABS 10MG /+10 30': {'UNITS': [0.0, 0.0, 0.0], 'PnVa': [0.0, 0.0, 0.0]}, u'ALORTIA TABS 50 MG /+5 30': {'UNITS': [0.0, 0.0, 0.0], 'PnVa': [0.0, 0.0, 0.0]}, u'ATORIS TOTAL': {'UNITS': [465.0, 402.0, 428.0], 'PnVa': [96.97, 101.97, 90.51]}}, u'A4.05': {u'ALORTIA TABS 100 MG /+5 30': {'UNITS': [0.0, 0.0, 0.0], 'PnVa': [0.0, 0.0, 0.0]}, u'ACLEXA TOTAL': {'UNITS': [3.0, 9.0, 62.0], 'PnVa': [6.58, 16.95, 48.04]}, u'SOBYCOMBI TABS 5MG /+10 30': {'UNITS': [-2.0, 1.0, 1.0], 'PnVa': [-733.76, 96.06, 57.07]}, u'SOBYCOMBI TABS 5MG /+5 30': {'UNITS': [-1.0, 2.0, 6.0], 'PnVa': [-14.27, 23.87, 48.69]}, u'KARBIS TOTAL': {'UNITS': [54.0, 61.0, 82.0], 'PnVa': [56.68, 67.84, 86.51]}, u'KARBICOMBI TOTAL': {'UNITS': [41.0, 45.0, 43.0], 'PnVa': [73.29, 93.99, 70.13]}, u'GLICLADA TOTAL': {'UNITS': [86.0, 130.0, 131.0], 'PnVa': [68.56, 108.69, 90.44]}, u'SOBYCOMBI TABS 10MG /+5 30': {'UNITS': [0.0, 1.0, 0.0], 'PnVa': [0.0, 112.15, 0.0]}, u'ATORIS FILM C.TABS 30 MG 30': {'UNITS': [18.0, 21.0, 16.0], 'PnVa': [25.51, 34.66, 22.89]}, u'ROSWERA BRAND': {'UNITS': [897.0, 748.0, 966.0], 'PnVa': [100.96, 99.93, 109.05]}, u'ALORTIA TABS 50 MG /+10 30': {'UNITS': [-3.0, 0.0, 1.0], 'PnVa': [-576.63, 0.0, 101.3]}, u'ALORTIA TABS 100 MG /+10 30': {'UNITS': [None, None, None], 'PnVa': [None, None, None]}, u'SOBYCOMBI TOTAL': {'UNITS': [-3.0, 5.0, 7.0], 'PnVa': [-43.05, 50.91, 39.66]}, u'ATORIS FILM C.TABS 30 MG 60': {'UNITS': [4.0, 10.0, 11.0], 'PnVa': [19.34, 51.18, 46.73]}, u'CO-AMLESSA TOTAL': {'UNITS': [8.0, 7.0, 11.0], 'PnVa': [32.84, 28.54, 35.19]}, u'ALORTIA TOTAL': {'UNITS': [-6.0, 0.0, 2.0], 'PnVa': [-100.9, 0.0, 19.58]}, u'SOBYCOMBI TABS 10MG /+10 30': {'UNITS': [0.0, 1.0, 0.0], 'PnVa': [0.0, 85.42, 0.0]}, u'ALORTIA TABS 50 MG /+5 30': {'UNITS': [-3.0, 0.0, 1.0], 'PnVa': [-65.95, 0.0, 16.12]}, u'ATORIS TOTAL': {'UNITS': [1351.0, 1117.0, 1251.0], 'PnVa': [101.6, 103.31, 93.43]}}, u'A4.02': {u'ALORTIA TABS 100 MG /+5 30': {'UNITS': [None, None, None], 'PnVa': [None, None, None]}, u'ACLEXA TOTAL': {'UNITS': [8.0, 23.0, 61.0], 'PnVa': [32.83, 59.54, 93.82]}, u'SOBYCOMBI TABS 5MG /+10 30': {'UNITS': [-4.0, -1.0, 3.0], 'PnVa': [-4332.06, -223.94, 326.83]}, u'SOBYCOMBI TABS 5MG /+5 30': {'UNITS': [1.0, 3.0, 4.0], 'PnVa': [42.14, 85.68, 60.22]}, u'KARBIS TOTAL': {'UNITS': [47.0, 44.0, 49.0], 'PnVa': [83.7, 73.36, 75.05]}, u'KARBICOMBI TOTAL': {'UNITS': [28.0, 37.0, 21.0], 'PnVa': [67.74, 84.04, 42.57]}, u'GLICLADA TOTAL': {'UNITS': [53.0, 63.0, 58.0], 'PnVa': [69.91, 98.15, 96.35]}, u'SOBYCOMBI TABS 10MG /+5 30': {'UNITS': [None, None, None], 'PnVa': [None, None, None]}, u'ATORIS FILM C.TABS 30 MG 30': {'UNITS': [4.0, 17.0, 5.0], 'PnVa': [11.65, 56.54, 15.32]}, u'ROSWERA BRAND': {'UNITS': [470.0, 551.0, 540.0], 'PnVa': [107.7, 138.7, 126.84]}, u'ALORTIA TABS 50 MG /+10 30': {'UNITS': [-3.0, 0.0, -1.0], 'PnVa': [-914.14, 0.0, -156.1]}, u'ALORTIA TABS 100 MG /+10 30': {'UNITS': [None, None, None], 'PnVa': [None, None, None]}, u'SOBYCOMBI TOTAL': {'UNITS': [-3.0, 2.0, 9.0], 'PnVa': [-161.48, 25.97, 126.73]}, u'ATORIS FILM C.TABS 30 MG 60': {'UNITS': [2.0, 7.0, -7.0], 'PnVa': [19.83, 72.13, -62.66]}, u'CO-AMLESSA TOTAL': {'UNITS': [6.0, 3.0, 3.0], 'PnVa': [44.89, 22.66, 17.92]}, u'ALORTIA TOTAL': {'UNITS': [-6.0, 0.0, -2.0], 'PnVa': [-161.42, 0.0, -33.2]}, u'SOBYCOMBI TABS 10MG /+10 30': {'UNITS': [0.0, 0.0, 2.0], 'PnVa': [0.0, 0.0, 301.53]}, u'ALORTIA TABS 50 MG /+5 30': {'UNITS': [-3.0, 0.0, -1.0], 'PnVa': [-104.55, 0.0, -26.02]}, u'ATORIS TOTAL': {'UNITS': [667.0, 631.0, 710.0], 'PnVa': [106.75, 117.04, 116.23]}}, u'A4.03': {u'ALORTIA TABS 100 MG /+5 30': {'UNITS': [None, None, None], 'PnVa': [None, None, None]}, u'ACLEXA TOTAL': {'UNITS': [4.0, 9.0, 34.0], 'PnVa': [7.05, 20.41, 57.52]}, u'SOBYCOMBI TABS 5MG /+10 30': {'UNITS': [-3.0, 0.0, 0.0], 'PnVa': [-1527.91, 0.0, 0.0]}, u'SOBYCOMBI TABS 5MG /+5 30': {'UNITS': [1.0, 2.0, 4.0], 'PnVa': [25.22, 33.87, 46.29]}, u'KARBIS TOTAL': {'UNITS': [33.0, 40.0, 57.0], 'PnVa': [44.81, 63.72, 78.75]}, u'KARBICOMBI TOTAL': {'UNITS': [12.0, 18.0, 16.0], 'PnVa': [33.03, 39.42, 29.6]}, u'GLICLADA TOTAL': {'UNITS': [67.0, 111.0, 96.0], 'PnVa': [87.25, 158.62, 103.11]}, u'SOBYCOMBI TABS 10MG /+5 30': {'UNITS': [0.0, 0.0, 0.0], 'PnVa': [0.0, 0.0, 0.0]}, u'ATORIS FILM C.TABS 30 MG 30': {'UNITS': [15.0, 13.0, 13.0], 'PnVa': [34.29, 34.47, 30.04]}, u'ROSWERA BRAND': {'UNITS': [525.0, 466.0, 547.0], 'PnVa': [94.16, 99.65, 97.15]}, u'ALORTIA TABS 50 MG /+10 30': {'UNITS': [0.0, 2.0, 0.0], 'PnVa': [0.0, 66.55, 7.88]}, u'ALORTIA TABS 100 MG /+10 30': {'UNITS': [None, None, None], 'PnVa': [None, None, None]}, u'SOBYCOMBI TOTAL': {'UNITS': [-2.0, 2.0, 5.0], 'PnVa': [-52.29, 21.14, 45.11]}, u'ATORIS FILM C.TABS 30 MG 60': {'UNITS': [3.0, 5.0, 4.0], 'PnVa': [23.21, 41.27, 27.33]}, u'CO-AMLESSA TOTAL': {'UNITS': [4.0, 7.0, 12.0], 'PnVa': [40.53, 42.49, 77.0]}, u'ALORTIA TOTAL': {'UNITS': [0.0, 2.0, -2.0], 'PnVa': [0.0, 18.58, -20.74]}, u'SOBYCOMBI TABS 10MG /+10 30': {'UNITS': [0.0, 0.0, 1.0], 'PnVa': [0.0, 0.0, 115.87]}, u'ALORTIA TABS 50 MG /+5 30': {'UNITS': [0.0, 0.0, -2.0], 'PnVa': [0.0, 0.0, -36.76]}, u'ATORIS TOTAL': {'UNITS': [741.0, 626.0, 722.0], 'PnVa': [89.07, 91.37, 88.49]}}, u'A4.01': {u'ALORTIA TABS 100 MG /+5 30': {'UNITS': [None, None, None], 'PnVa': [None, None, None]}, u'ACLEXA TOTAL': {'UNITS': [0.0, 2.0, 34.0], 'PnVa': [8.32, 13.99, 125.82]}, u'SOBYCOMBI TABS 5MG /+10 30': {'UNITS': [0.0, 0.0, 0.0], 'PnVa': [0.0, 0.0, 0.0]}, u'SOBYCOMBI TABS 5MG /+5 30': {'UNITS': [0.0, -1.0, 0.0], 'PnVa': [0.0, -44.49, 0.0]}, u'KARBIS TOTAL': {'UNITS': [25.0, 5.0, 13.0], 'PnVa': [140.91, 29.27, 95.62]}, u'KARBICOMBI TOTAL': {'UNITS': [18.0, 14.0, 9.0], 'PnVa': [135.11, 105.07, 84.11]}, u'GLICLADA TOTAL': {'UNITS': [25.0, 23.0, 27.0], 'PnVa': [50.67, 82.99, 77.75]}, u'SOBYCOMBI TABS 10MG /+5 30': {'UNITS': [0.0, 0.0, 0.0], 'PnVa': [0.0, 0.0, 0.0]}, u'ATORIS FILM C.TABS 30 MG 30': {'UNITS': [7.0, 12.0, 13.0], 'PnVa': [46.53, 98.0, 94.47]}, u'ROSWERA BRAND': {'UNITS': [197.0, 199.0, 167.0], 'PnVa': [108.9, 137.62, 98.92]}, u'ALORTIA TABS 50 MG /+10 30': {'UNITS': [0.0, 0.0, 0.0], 'PnVa': [0.0, 0.0, 0.0]}, u'ALORTIA TABS 100 MG /+10 30': {'UNITS': [None, None, None], 'PnVa': [None, None, None]}, u'SOBYCOMBI TOTAL': {'UNITS': [0.0, -1.0, 0.0], 'PnVa': [0.0, -27.76, 0.0]}, u'ATORIS FILM C.TABS 30 MG 60': {'UNITS': [2.0, 0.0, 3.0], 'PnVa': [45.15, 0.0, 64.46]}, u'CO-AMLESSA TOTAL': {'UNITS': [4.0, 3.0, 4.0], 'PnVa': [54.73, 39.69, 52.32]}, u'ALORTIA TOTAL': {'UNITS': [0.0, 0.0, 1.0], 'PnVa': [0.0, 0.0, 41.66]}, u'SOBYCOMBI TABS 10MG /+10 30': {'UNITS': [0.0, 0.0, 0.0], 'PnVa': [0.0, 0.0, 0.0]}, u'ALORTIA TABS 50 MG /+5 30': {'UNITS': [0.0, 0.0, 1.0], 'PnVa': [0.0, 0.0, 77.19]}, u'ATORIS TOTAL': {'UNITS': [266.0, 178.0, 236.0], 'PnVa': [102.07, 72.37, 96.5]}}, u'A4.08': {u'ALORTIA TABS 100 MG /+5 30': {'UNITS': [None, None, None], 'PnVa': [None, None, None]}, u'ACLEXA TOTAL': {'UNITS': [8.0, 12.0, 51.0], 'PnVa': [11.18, 10.0, 31.08]}, u'SOBYCOMBI TABS 5MG /+10 30': {'UNITS': [0.0, -4.0, 3.0], 'PnVa': [15.72, -383.77, 113.5]}, u'SOBYCOMBI TABS 5MG /+5 30': {'UNITS': [6.0, 4.0, 14.0], 'PnVa': [72.04, 51.39, 75.43]}, u'KARBIS TOTAL': {'UNITS': [99.0, 115.0, 91.0], 'PnVa': [83.98, 103.45, 69.15]}, u'KARBICOMBI TOTAL': {'UNITS': [65.0, 40.0, 64.0], 'PnVa': [110.17, 70.57, 113.0]}, u'GLICLADA TOTAL': {'UNITS': [140.0, 123.0, 109.0], 'PnVa': [78.31, 82.54, 60.33]}, u'SOBYCOMBI TABS 10MG /+5 30': {'UNITS': [0.0, 0.0, 1.0], 'PnVa': [0.0, 0.0, 66.01]}, u'ATORIS FILM C.TABS 30 MG 30': {'UNITS': [52.0, 33.0, 43.0], 'PnVa': [56.55, 43.63, 50.21]}, u'ROSWERA BRAND': {'UNITS': [1292.0, 1069.0, 1299.0], 'PnVa': [120.18, 118.63, 123.61]}, u'ALORTIA TABS 50 MG /+10 30': {'UNITS': [0.0, 2.0, -2.0], 'PnVa': [0.0, 47.43, -142.45]}, u'ALORTIA TABS 100 MG /+10 30': {'UNITS': [0.0, 0.0, 0.0], 'PnVa': [0.0, 0.0, 0.0]}, u'SOBYCOMBI TOTAL': {'UNITS': [6.0, 1.0, 18.0], 'PnVa': [49.09, -0.65, 70.96]}, u'ATORIS FILM C.TABS 30 MG 60': {'UNITS': [9.0, 14.0, 13.0], 'PnVa': [33.23, 57.44, 44.91]}, u'CO-AMLESSA TOTAL': {'UNITS': [5.0, 2.0, 17.0], 'PnVa': [18.74, 6.85, 51.9]}, u'ALORTIA TOTAL': {'UNITS': [0.0, 3.0, -5.0], 'PnVa': [0.0, 18.39, -37.46]}, u'SOBYCOMBI TABS 10MG /+10 30': {'UNITS': [0.0, 1.0, 0.0], 'PnVa': [0.0, 90.05, 0.0]}, u'ALORTIA TABS 50 MG /+5 30': {'UNITS': [0.0, 1.0, -3.0], 'PnVa': [0.0, 11.16, -36.94]}, u'ATORIS TOTAL': {'UNITS': [1649.0, 1362.0, 1399.0], 'PnVa': [95.55, 96.22, 91.68]}}, u'A4': {u'ALORTIA TABS 100 MG /+5 30': {'UNITS': [0.0, 0.0, 0.0], 'PnVa': [0.0, 0.0, 0.0]}, u'ACLEXA TOTAL': {'UNITS': [44.0, 169.0, 390.0], 'PnVa': [12.85, 37.34, 60.46]}, u'SOBYCOMBI TABS 5MG /+10 30': {'UNITS': [-10.0, -1.0, 5.0], 'PnVa': [-743.84, -15.43, 55.76]}, u'SOBYCOMBI TABS 5MG /+5 30': {'UNITS': [6.0, 9.0, 34.0], 'PnVa': [20.22, 23.85, 52.11]}, u'KARBIS TOTAL': {'UNITS': [340.0, 335.0, 376.0], 'PnVa': [65.2, 68.96, 70.47]}, u'KARBICOMBI TOTAL': {'UNITS': [223.0, 207.0, 206.0], 'PnVa': [70.82, 64.61, 61.49]}, u'GLICLADA TOTAL': {'UNITS': [510.0, 599.0, 596.0], 'PnVa': [66.79, 93.53, 80.13]}, u'SOBYCOMBI TABS 10MG /+5 30': {'UNITS': [1.0, 0.0, 3.0], 'PnVa': [26.85, 1.53, 56.17]}, u'ATORIS FILM C.TABS 30 MG 30': {'UNITS': [171.0, 175.0, 141.0], 'PnVa': [44.22, 53.97, 38.45]}, u'ROSWERA BRAND': {'UNITS': [5122.0, 4442.0, 5279.0], 'PnVa': [107.7, 110.04, 114.31]}, u'ALORTIA TABS 50 MG /+10 30': {'UNITS': [-7.0, 4.0, -3.0], 'PnVa': [-228.81, 18.88, -41.51]}, u'ALORTIA TABS 100 MG /+10 30': {'UNITS': [0.0, 0.0, 2.0], 'PnVa': [0.0, 0.0, 32.54]}, u'SOBYCOMBI TOTAL': {'UNITS': [-2.0, 11.0, 46.0], 'PnVa': [-12.5, 22.47, 53.89]}, u'ATORIS FILM C.TABS 30 MG 60': {'UNITS': [36.0, 49.0, 43.0], 'PnVa': [31.61, 46.83, 34.71]}, u'CO-AMLESSA TOTAL': {'UNITS': [50.0, 30.0, 62.0], 'PnVa': [41.34, 24.46, 41.29]}, u'ALORTIA TOTAL': {'UNITS': [-14.0, 6.0, -6.0], 'PnVa': [-40.25, 7.35, -6.63]}, u'SOBYCOMBI TABS 10MG /+10 30': {'UNITS': [1.0, 3.0, 4.0], 'PnVa': [25.76, 56.67, 60.46]}, u'ALORTIA TABS 50 MG /+5 30': {'UNITS': [-7.0, 2.0, -5.0], 'PnVa': [-26.17, 4.35, -12.28]}, u'ATORIS TOTAL': {'UNITS': [6866.0, 5698.0, 6390.0], 'PnVa': [97.28, 98.7, 94.35]}}}

    def start_output(self):
        """inicjalizacja pliku output i tworzenie w nim zakłądek"""

        self.output_file = openpyxl.Workbook()
        for elem in range(len(self.output_lista_cegiel)):
            self.output_file.create_sheet(elem, self.output_lista_cegiel[elem])

        # usuwa pierwszą z stworzonych domyślną zakładkę - pewnie można to jakoś obejść .....
        self.output_file.remove_sheet(self.output_file.get_sheet_by_name('Sheet'))
        self.output_sheets = self.output_file.get_sheet_names()

    def save_output(self):
        """zapisuje dane dla pliku output w podanej lokalizacji"""

        self.output_file.save('test_excel_file_CEGLY.xlsx')

    def write_base_data(self):
        """Zapisuje podstawowe dane w pliku output UNITS"""

        for elem in range(len(self.output_lista_cegiel)):

            sheet = self.output_lista_cegiel[elem]
            sh = self.output_file.get_sheet_by_name(sheet)
            col = 1

            sh['B3'] = "UNITS"
            sh['E3'] = "PnVa"
            for i in range(2):
                for ii in range(3):
                    sh[self.alfabet[col] + '4'] = self.daty[ii]
                    col += 1
            col, row = 0, 5
            for i in self.output_leki_cegly:
                sh[self.alfabet[col] + str(row)] = i
                row += 1

    def write_PnVa_UNITS_data(self):
        """Metoda zapisuje w pliku output dany PnVa oraz UNITS"""

        for elem in range(len(self.output_lista_cegiel)):

            col, row = 1, 5
            sheet = self.output_lista_cegiel[elem]
            sh = self.output_file.get_sheet_by_name(sheet)

            for i in self.output_leki_cegly:
                for x in range(3):
                    sh[self.alfabet[col] + str(row)] = self.CEGLY_data[sheet][i]["UNITS"][x]
                    col += 1
                for x in range(3):
                    sh[self.alfabet[col] + str(row)] = self.CEGLY_data[sheet][i]["PnVa"][x]
                    col += 1
                col = 1
                row += 1




if __name__ == '__main__':
    print 50 * "xxx"
    print "output1"

    input1 = Read_input(input_file_path1)
    input1.get_Cegly_data()

    print "input1.CEGLY_data", input1.CEGLY_data

    output1 = Write_output_PnVa_Units_paternA("PnVa")
    output1.start_output()
    output1.write_base_data()
    output1.write_PnVa_UNITS_data()
    output1.save_output()

    output2 = Write_output_CEGLY_paternA()
    output2.start_output()
    output2.write_base_data()
    output2.write_PnVa_UNITS_data()
    output2.save_output()




